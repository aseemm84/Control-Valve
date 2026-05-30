"""
PDF and Excel report generation.
==================================
Builds downloadable engineering reports from a SizingResult.
Returns BytesIO objects consumed by st.download_button().
"""

from __future__ import annotations

import io
from datetime import datetime
from typing import Optional

import streamlit as st

from backend.models import FluidPhase, SizingResult


# ---------------------------------------------------------------------------
# UNICODE TO ASCII SAFETY FILTER
# ---------------------------------------------------------------------------

def _pdf_safe(text: str) -> str:
    """Replace Unicode chars outside Latin-1 with ASCII engineering equivalents."""
    if not isinstance(text, str):
        text = str(text)

    replacements = {
        "\u2014": "-",      "\u2013": "-",      "\u2212": "-",
        "\u0394": "Delta",  "\u03c1": "rho",    "\u03c3": "sigma",
        "\u03b7": "eta",    "\u03b3": "gamma",  "\u03bc": "mu",
        "\u03bd": "nu",     "\u03c6": "phi",    "\u03c0": "pi",
        "\u03b1": "alpha",  "\u03b2": "beta",   "\u03b5": "eps",
        "\u03ba": "kappa",  "\u03c9": "omega",  "\u03a9": "Omega",
        "\u03a3": "Sigma",
        "\u2080": "0",      "\u2081": "1",      "\u2082": "2",
        "\u2083": "3",      "\u2084": "4",
        "\u00b2": "2",      "\u00b3": "3",      "\u00b9": "1",
        "\u00d7": "x",      "\u00b7": ".",      "\u221a": "sqrt",
        "\u2265": ">=",     "\u2264": "<=",     "\u00b1": "+/-",
        "\u221e": "inf",    "\u2192": "->",     "\u00b0": "deg",
        "\u2018": "'",      "\u2019": "'",
        "\u201c": '"',      "\u201d": '"',
        "\u2022": "*",
    }
    for char, replacement in replacements.items():
        text = text.replace(char, replacement)
    return text.encode("latin-1", errors="replace").decode("latin-1")


def _fmt(value: Optional[float], decimals: int = 3) -> str:
    """Format a float for PDF output. Returns '-' when None."""
    if value is None:
        return "-"
    return f"{value:.{decimals}f}"


# ---------------------------------------------------------------------------
# MATPLOTLIB CHART GENERATORS  (return BytesIO PNG or None on failure)
# ---------------------------------------------------------------------------

def _mpl():
    """Import matplotlib with Agg (non-GUI) backend."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    plt.rcParams.update({
        "font.family":    "sans-serif",
        "font.size":      8,
        "axes.titlesize": 9,
        "axes.labelsize": 8,
        "axes.grid":      True,
        "grid.alpha":     0.3,
        "figure.dpi":     150,
    })
    return plt


def _save_png(fig) -> io.BytesIO:
    """Save a matplotlib figure to a BytesIO PNG buffer."""
    import matplotlib.pyplot as plt
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight",
                facecolor="white")
    plt.close(fig)
    buf.seek(0)
    return buf


def _chart_cv_curve(
    Cv_rated: float,
    Cv_required: float,
    opening_pct: Optional[float],
    R_inherent: float = 50.0,
) -> Optional[io.BytesIO]:
    """Cv characteristic curve with operating point marker."""
    try:
        import numpy as np
        plt = _mpl()
        fig, ax = plt.subplots(figsize=(6.0, 3.0))

        x    = np.linspace(0, 100, 201)
        Cmin = Cv_rated / max(R_inherent, 1.0)

        # Three curves
        ep  = Cmin * (Cv_rated / Cmin) ** (x / 100)
        lin = Cmin + (Cv_rated - Cmin) * x / 100
        qo  = np.where(x > 0, Cv_rated * np.sqrt(x / 100), 0)

        ax.plot(x, ep,  color="#1B6CA8", lw=2.0, label="Equal Percentage")
        ax.plot(x, lin, color="#198754", lw=1.2, ls="--", alpha=0.7, label="Linear")
        ax.plot(x, qo,  color="#fd7e14", lw=1.2, ls=":",  alpha=0.7, label="Quick Opening")

        # Controllable zone band
        ax.axvspan(10, 90, alpha=0.06, color="#198754", zorder=0)
        ax.text(50, Cv_rated * 0.04, "Controllable zone (10-90%)",
                ha="center", va="bottom", fontsize=6.5, color="#198754")

        # Operating point
        if Cv_required and opening_pct is not None:
            ax.scatter([opening_pct], [Cv_required],
                       color="#dc3545", s=70, zorder=5,
                       marker="D", label=f"Op. Point  Cv={Cv_required:.1f}")
            ax.annotate(
                f"  Cv={Cv_required:.1f}",
                xy=(opening_pct, Cv_required),
                fontsize=7.5, color="#dc3545", va="center",
            )

        ax.set_xlim(0, 100)
        ax.set_ylim(0, Cv_rated * 1.08)
        ax.set_xlabel("Valve Opening [%]")
        ax.set_ylabel("Cv [-]")
        ax.set_title("Inherent Flow Characteristic Curves")
        ax.legend(loc="upper left", fontsize=7, framealpha=0.8)
        fig.tight_layout()
        return _save_png(fig)
    except Exception:
        return None


def _chart_sizing_gauge(
    sizing_ratio: float,
    Cv_required: float,
    Cv_rated: float,
) -> Optional[io.BytesIO]:
    """Horizontal gauge bar showing sizing ratio with zone colouring."""
    try:
        plt = _mpl()
        fig, ax = plt.subplots(figsize=(5.0, 1.4))

        pct = min(sizing_ratio * 100, 120)

        # Zone backgrounds
        zones = [
            (0,   20,  "#fde8e8"),   # red: too small
            (20,  60,  "#fff3cd"),   # yellow: oversized
            (60,  85,  "#d4edda"),   # green: optimal
            (85,  100, "#fff3cd"),   # yellow: near capacity
            (100, 120, "#fde8e8"),   # red: undersized
        ]
        for lo, hi, col in zones:
            ax.barh(0, hi - lo, left=lo, height=0.55,
                    color=col, alpha=0.9, zorder=1)

        # Value bar
        bar_col = (
            "#198754" if 60 <= pct <= 85
            else "#fd7e14" if 20 < pct < 100
            else "#dc3545"
        )
        ax.barh(0, pct, height=0.35, color=bar_col, alpha=0.95, zorder=2)

        # Zone boundary lines
        for v, lbl in [(60, "60%"), (85, "85%"), (100, "100%")]:
            ax.axvline(v, color="gray", lw=0.8, ls="--", zorder=3)
            ax.text(v, 0.42, lbl, ha="center", va="bottom",
                    fontsize=6, color="gray")

        ax.set_xlim(0, 120)
        ax.set_ylim(-0.5, 0.7)
        ax.set_yticks([])
        ax.set_xlabel("Sizing Ratio [%]")
        ax.set_title(
            f"Sizing Ratio: {pct:.1f}%   "
            f"(Cv {Cv_required:.1f} / {Cv_rated:.1f})   "
            f"Optimal: 60-85%",
            fontsize=8,
        )
        fig.tight_layout()
        return _save_png(fig)
    except Exception:
        return None


def _chart_pressure_profile(
    P1: float,
    P_vc: float,
    P2: float,
    Pv: float,
) -> Optional[io.BytesIO]:
    """Horizontal bar chart showing P1, Pvc, P2 vs Pv threshold."""
    try:
        plt = _mpl()
        fig, ax = plt.subplots(figsize=(6.5, 2.0))

        labels  = ["P1 (Inlet)", "Pvc (Vena Contracta)", "P2 (Outlet)"]
        values  = [max(v, 0.0) for v in [P1, P_vc, P2]]
        colours = ["#1B6CA8", "#fd7e14", "#198754"]

        bars = ax.barh(labels, values, color=colours, alpha=0.85, height=0.5)

        for bar, raw_val in zip(bars, [P1, P_vc, P2]):
            bw = bar.get_width()
            ax.text(
                bw * 0.5, bar.get_y() + bar.get_height() / 2,
                f"{raw_val:.3f} bar",
                ha="center", va="center",
                color="white", fontweight="bold", fontsize=8,
            )

        ax.axvline(Pv, color="#dc3545", ls="--", lw=1.5,
                   label=f"Pv = {Pv:.4f} bar")

        ax.set_xlabel("Absolute Pressure [bar]")
        ax.set_title("Pressure Profile Through Valve")
        ax.legend(fontsize=7.5, loc="lower right")
        ax.set_xlim(left=0)
        fig.tight_layout()
        return _save_png(fig)
    except Exception:
        return None


# ---------------------------------------------------------------------------
# CHART PAGE EMBEDDER
# ---------------------------------------------------------------------------

def _embed_chart_page(pdf, result: SizingResult) -> None:
    """Add a CHARTS section to the PDF with embedded PNG images.

    Embeds up to three charts:
      1. Cv Characteristic Curve     (always, when Cv_rated can be derived)
      2. Sizing Ratio Gauge          (always, when sizing_ratio available)
      3. Pressure Profile Through Valve  (liquid only)

    Each chart is wrapped in try-except so a single failure never
    prevents the rest of the PDF from generating.
    """
    has_sizing = (
        result.Cv_required is not None
        and result.sizing_ratio is not None
        and result.sizing_ratio > 0
    )

    if not has_sizing and result.fluid_phase != FluidPhase.LIQUID:
        return   # nothing to draw

    pdf.section_title("CHARTS")

    if has_sizing:
        Cv_rated_est = result.Cv_required / result.sizing_ratio

        # ── Row 1: Cv curve (left 64%) + Sizing gauge (right 34%) ────────
        y0     = pdf.get_y()
        w_cv   = pdf.epw * 0.64
        w_gau  = pdf.epw * 0.34
        gap    = pdf.epw - w_cv - w_gau   # 2% gap

        cv_buf = _chart_cv_curve(
            Cv_rated_est,
            result.Cv_required,
            result.opening_pct,
        )
        if cv_buf:
            # aspect of figsize(6, 3) → height = width × (3/6)
            h_cv = w_cv * (3.0 / 6.0)
            pdf.image(
                cv_buf,
                x=pdf.l_margin,
                y=y0,
                w=w_cv,
            )
        else:
            h_cv = 0

        ga_buf = _chart_sizing_gauge(
            result.sizing_ratio,
            result.Cv_required,
            Cv_rated_est,
        )
        if ga_buf:
            # aspect of figsize(5, 1.4) → height = width × (1.4/5)
            h_ga = w_gau * (1.4 / 5.0)
            pdf.image(
                ga_buf,
                x=pdf.l_margin + w_cv + gap,
                y=y0,
                w=w_gau,
            )
        else:
            h_ga = 0

        # Move y past the taller of the two images
        pdf.set_y(y0 + max(h_cv, h_ga) + 5)

    # ── Row 2: Pressure profile (liquid, full width) ──────────────────────
    if (
        result.fluid_phase == FluidPhase.LIQUID
        and result.cavitation is not None
        and result.P1_bar is not None
        and result.P2_bar is not None
    ):
        cav = result.cavitation
        # Derive Pv from sigma: sigma = (P1-Pv)/(P1-P2)
        delta_P = result.P1_bar - result.P2_bar
        Pv_est  = result.P1_bar - cav.sigma * delta_P if delta_P > 0 else 0.0

        pr_buf = _chart_pressure_profile(
            result.P1_bar,
            cav.P_vc,
            result.P2_bar,
            Pv_est,
        )
        if pr_buf:
            pdf.set_x(pdf.l_margin)
            # aspect of figsize(6.5, 2) → height = width × (2/6.5)
            pdf.image(pr_buf, x=pdf.l_margin, w=pdf.epw)
            pdf.ln(5)

    pdf.ln(2)


# ---------------------------------------------------------------------------
# PDF REPORT
# ---------------------------------------------------------------------------

def build_pdf_bytes(result: SizingResult) -> bytes:
    """Generate an engineering sizing report as PDF bytes.

    Returns
    -------
    bytes
        Raw PDF bytes for st.download_button().
    """
    try:
        from fpdf import FPDF
    except ImportError:
        raise ImportError("fpdf2 is required. Run: pip install fpdf2>=2.7.6")

    class _Report(FPDF):

        def header(self) -> None:
            self.set_fill_color(27, 108, 168)
            self.rect(0, 0, 210, 16, "F")
            self.set_font("Helvetica", "B", 13)
            self.set_text_color(255, 255, 255)
            self.set_x(self.l_margin)
            self.cell(
                self.epw, 16,
                "Control Valve Sizing Report",
                align="C",
                new_x="LMARGIN",
                new_y="NEXT",
            )
            self.set_text_color(0, 0, 0)

        def footer(self) -> None:
            self.set_y(-12)
            self.set_x(self.l_margin)
            self.set_font("Helvetica", "I", 8)
            self.set_text_color(120, 120, 120)
            self.cell(
                self.epw, 10,
                (
                    f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} UTC"
                    "   |   IEC 60534-2-1:2011 / ISA-75.01.01-2012"
                    f"   |   Page {self.page_no()}"
                ),
                align="C",
                new_x="LMARGIN",
                new_y="NEXT",
            )

        def section_title(self, title: str) -> None:
            self.set_x(self.l_margin)
            self.set_fill_color(27, 108, 168)
            self.set_text_color(255, 255, 255)
            self.set_font("Helvetica", "B", 10)
            self.cell(
                self.epw, 7,
                f"  {_pdf_safe(title)}",
                fill=True,
                new_x="LMARGIN",
                new_y="NEXT",
            )
            self.set_text_color(0, 0, 0)
            self.ln(2)

        def data_row(self, label: str, value: str, unit: str = "") -> None:
            self.set_x(self.l_margin)
            epw  = self.epw
            col1 = epw * 0.44
            col2 = epw * 0.34
            col3 = epw - col1 - col2

            even = (int(self.get_y()) // 6) % 2 == 0
            self.set_fill_color(248, 249, 250)

            self.set_font("Helvetica", "", 9)
            self.set_text_color(0, 0, 0)
            self.cell(col1, 6, _pdf_safe(label), border="B", fill=even)

            self.set_font("Helvetica", "B", 9)
            self.set_text_color(27, 108, 168)
            self.cell(col2, 6, _pdf_safe(value), border="B", fill=even)

            self.set_text_color(100, 100, 100)
            self.set_font("Helvetica", "I", 9)
            self.cell(
                col3, 6, _pdf_safe(unit),
                border="B", fill=even,
                new_x="LMARGIN", new_y="NEXT",
            )
            self.set_text_color(0, 0, 0)

    # ── Build document ─────────────────────────────────────────────────────
    pdf = _Report()
    pdf.set_margins(left=12, top=15, right=12)
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Subtitle
    pdf.set_x(pdf.l_margin)
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(
        pdf.epw, 6,
        f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  Revision: A  |  Status: Preliminary",
        align="C",
        new_x="LMARGIN",
        new_y="NEXT",
    )
    pdf.set_text_color(0, 0, 0)
    pdf.ln(4)

    # ── Section 1: Primary Sizing Results ─────────────────────────────────
    pdf.section_title("1. PRIMARY SIZING RESULTS")
    pdf.data_row("Cv Required",             _fmt(result.Cv_required),      "[-]")
    pdf.data_row("Cv Design (with margin)", _fmt(result.Cv_design),        "[-]")
    pdf.data_row("Kv Required",             _fmt(result.Kv_required),      "m3/h/bar^0.5")
    if result.sizing_ratio is not None:
        pdf.data_row("Sizing Ratio",
                     f"{result.sizing_ratio * 100:.1f}%",                  "[-]")
    if result.opening_pct is not None:
        pdf.data_row("Estimated Opening",
                     f"{result.opening_pct:.1f}%",                         "%")
    if result.velocity_ms is not None:
        pdf.data_row("Downstream Velocity",
                     _fmt(result.velocity_ms, 2),                          "m/s")
    pdf.ln(4)

    # ── Section 2: Process Conditions ─────────────────────────────────────
    pdf.section_title("2. PROCESS CONDITIONS (SI Internal Values)")
    pdf.data_row("Inlet Pressure P1",       _fmt(result.P1_bar, 4),        "bar abs")
    pdf.data_row("Outlet Pressure P2",      _fmt(result.P2_bar, 4),        "bar abs")
    if result.P1_bar and result.P2_bar:
        pdf.data_row("Differential dP",
                     _fmt(result.P1_bar - result.P2_bar, 4),               "bar")
    pdf.data_row("Inlet Temperature T1",
                 _fmt(result.T1_K - 273.15 if result.T1_K else None, 1),   "deg C")
    pdf.data_row("Mass Flow W",             _fmt(result.W_kgh, 3),         "kg/h")
    pdf.data_row("Inlet Density rho1",      _fmt(result.rho1_kgm3, 4),     "kg/m3")
    pdf.ln(4)

    # ── Section 3: Intermediate Values ────────────────────────────────────
    pdf.section_title("3. INTERMEDIATE COMPUTED VALUES")
    pdf.data_row("Fluid Phase",
                 result.fluid_phase.value if result.fluid_phase else "-",  "")
    pdf.data_row("Piping Factor Fp",        _fmt(result.Fp, 4),            "[-]")
    pdf.data_row("FF Factor",               _fmt(result.FF, 4),            "[-]")
    pdf.data_row("dP max (choked)",         _fmt(result.delta_P_max_bar, 4), "bar")
    pdf.data_row("dP eff (used in Cv)",     _fmt(result.delta_P_eff_bar, 4), "bar")
    pdf.data_row("Expansion Factor Y",      _fmt(result.Y, 4),             "[-]")
    pdf.data_row("Fgamma",                  _fmt(result.Fgamma, 4),        "[-]")
    pdf.data_row("x = dP/P1",              _fmt(result.x, 4),             "[-]")
    pdf.data_row("Reynolds Number Rev",
                 _fmt(result.Rev, 0) if result.Rev else "-",               "[-]")
    pdf.data_row("FR Factor",               _fmt(result.FR, 4),            "[-]")
    if result.FLP is not None:
        pdf.data_row("FLP (with fittings)", _fmt(result.FLP, 4),           "[-]")
    if result.xTP is not None:
        pdf.data_row("xTP (with fittings)", _fmt(result.xTP, 4),           "[-]")
    pdf.ln(4)

    # ── Section 4: Cavitation ─────────────────────────────────────────────
    if result.cavitation:
        cav = result.cavitation
        pdf.section_title("4. CAVITATION / FLASHING ANALYSIS (IEC 60534-8-4)")
        pdf.data_row("Cavitation Regime",   cav.regime.value.upper(),      "")
        pdf.data_row("Sigma",               _fmt(cav.sigma, 4),            "[-]")
        pdf.data_row("Vena Contracta Pvc",  _fmt(cav.P_vc, 4),            "bar abs")
        pdf.data_row("dP Incipient",        _fmt(cav.delta_P_incipient, 4), "bar")
        pdf.data_row("dP Max",              _fmt(cav.delta_P_max, 4),      "bar")
        pdf.data_row("FF Factor",           _fmt(cav.FF, 4),               "[-]")
        pdf.data_row("Is Choked",           str(cav.is_choked),            "")
        pdf.data_row("Is Flashing",         str(cav.is_flashing),          "")
        pdf.ln(4)

    # ── Section 5: Noise ──────────────────────────────────────────────────
    if result.noise:
        noise = result.noise
        pdf.section_title("5. NOISE PREDICTION (IEC 60534-8-3 / 8-4)")
        pdf.data_row("External SPL Lpe",    f"{noise.Lpe_dba:.1f}",        "dBA at 1 m")
        pdf.data_row("Internal LWi",        f"{noise.LWi_db:.1f}",         "dB re 1 pW")
        pdf.data_row("Pipe Wall TL",        f"{noise.TL_db:.1f}",          "dB")
        pdf.data_row("Peak Frequency f_p",  f"{noise.f_peak_hz:.0f}",      "Hz")
        pdf.data_row("Acoustic Eff. eta",   f"{noise.eta:.2e}",            "[-]")
        pdf.data_row("Flow Regime",         noise.regime,                  "")
        pdf.data_row("Exceeds Site Limit",  str(noise.exceeds_limit),      "")
        pdf.ln(4)

    # ── Section 6: Engineering Warnings ───────────────────────────────────
    if result.messages:
        pdf.section_title("6. ENGINEERING WARNINGS AND NOTES")
        for msg in result.messages:
            pdf.set_x(pdf.l_margin)
            level_str = f"[{msg.level.value.upper()}]"

            if msg.level.value == "error":
                pdf.set_text_color(220, 53, 69)
            elif msg.level.value == "warning":
                pdf.set_text_color(200, 100, 4)
            else:
                pdf.set_text_color(13, 110, 253)

            pdf.set_font("Helvetica", "B", 9)
            pdf.cell(32, 6, level_str)

            pdf.set_text_color(0, 0, 0)
            pdf.set_font("Helvetica", "", 9)
            pdf.multi_cell(
                pdf.epw - 32, 6,
                f"{_pdf_safe(msg.code)}  {_pdf_safe(msg.message)}",
                new_x="LMARGIN",
                new_y="NEXT",
            )

    # ── CHARTS (new page for clean layout) ────────────────────────────────
    pdf.add_page()
    _embed_chart_page(pdf, result)

    return bytes(pdf.output())


# ---------------------------------------------------------------------------
# EXCEL REPORT
# ---------------------------------------------------------------------------

def build_excel_bytes(result: SizingResult) -> bytes:
    """Generate a structured Excel workbook with sizing results."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError:
        raise ImportError("openpyxl is required.")

    wb = openpyxl.Workbook()
    hdr_font    = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill    = PatternFill("solid", fgColor="1B6CA8")
    val_font    = Font(bold=True, color="1B6CA8", size=10)
    lbl_font    = Font(size=10)
    thin_border = Border(bottom=Side(style="thin", color="DEE2E6"))
    center      = Alignment(horizontal="center", vertical="center")

    def write_section(ws, row: int, title: str) -> int:
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        return row + 1

    def write_row(ws, row: int, label: str, value: str, unit: str = "") -> int:
        ws.cell(row=row, column=1, value=label).font = lbl_font
        ws.cell(row=row, column=2, value=value).font = val_font
        ws.cell(row=row, column=3, value=unit).font  = Font(color="6C757D", italic=True)
        for col in range(1, 4):
            ws.cell(row=row, column=col).border = thin_border
        return row + 1

    ws1 = wb.active
    ws1.title = "Sizing Results"
    ws1.column_dimensions["A"].width = 38
    ws1.column_dimensions["B"].width = 22
    ws1.column_dimensions["C"].width = 22

    r = 1
    tc = ws1.cell(r, 1, "CONTROL VALVE SIZING REPORT")
    tc.font = Font(bold=True, size=14, color="1B6CA8")
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 1
    ws1.cell(r, 1,
             f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(
        color="6C757D", size=9)
    r += 2

    r = write_section(ws1, r, "PRIMARY RESULTS")
    r = write_row(ws1, r, "Cv Required",
                  _fmt(result.Cv_required),         "\u2014")
    r = write_row(ws1, r, "Cv Design (with margin)",
                  _fmt(result.Cv_design),            "\u2014")
    r = write_row(ws1, r, "Kv Required",
                  _fmt(result.Kv_required),          "m\u00b3/h/bar^0.5")
    r = write_row(ws1, r, "Sizing Ratio",
                  f"{result.sizing_ratio*100:.1f}%"
                  if result.sizing_ratio else "\u2014",  "\u2014")
    r = write_row(ws1, r, "Opening %",
                  f"{result.opening_pct:.1f}%"
                  if result.opening_pct else "\u2014",    "%")
    r = write_row(ws1, r, "Downstream Velocity",
                  _fmt(result.velocity_ms, 2)
                  if result.velocity_ms else "\u2014",    "m/s")
    r += 1

    r = write_section(ws1, r, "PROCESS CONDITIONS (SI)")
    r = write_row(ws1, r, "Inlet Pressure P1",
                  _fmt(result.P1_bar, 4),            "bar abs")
    r = write_row(ws1, r, "Outlet Pressure P2",
                  _fmt(result.P2_bar, 4),            "bar abs")
    r = write_row(ws1, r, "Inlet Temperature T1",
                  _fmt(result.T1_K - 273.15 if result.T1_K else None, 1),
                  "\u00b0C")
    r = write_row(ws1, r, "Mass Flow W",
                  _fmt(result.W_kgh, 3),             "kg/h")
    r = write_row(ws1, r, "Inlet Density \u03c1\u2081",
                  _fmt(result.rho1_kgm3, 4),         "kg/m\u00b3")
    r += 1

    r = write_section(ws1, r, "INTERMEDIATE VALUES")
    r = write_row(ws1, r, "Piping Factor Fp",   _fmt(result.Fp, 4),  "\u2014")
    r = write_row(ws1, r, "FF Factor",          _fmt(result.FF, 4),  "\u2014")
    r = write_row(ws1, r, "\u0394P max",        _fmt(result.delta_P_max_bar, 4), "bar")
    r = write_row(ws1, r, "Expansion Factor Y", _fmt(result.Y, 4),   "\u2014")
    r = write_row(ws1, r, "Reynolds Rev",
                  _fmt(result.Rev, 0) if result.Rev else "\u2014",   "\u2014")
    r = write_row(ws1, r, "FR Factor",          _fmt(result.FR, 4),  "\u2014")
    r += 1

    if result.cavitation:
        cav = result.cavitation
        ws2 = wb.create_sheet("Cavitation Analysis")
        ws2.column_dimensions["A"].width = 38
        ws2.column_dimensions["B"].width = 22
        ws2.column_dimensions["C"].width = 22
        r2 = write_section(ws2, 1, "CAVITATION ANALYSIS (IEC 60534-8-4)")
        r2 = write_row(ws2, r2, "Regime",    cav.regime.value.upper(), "")
        r2 = write_row(ws2, r2, "Sigma",     _fmt(cav.sigma, 4),       "\u2014")
        r2 = write_row(ws2, r2, "Pvc",       _fmt(cav.P_vc, 4),        "bar abs")
        r2 = write_row(ws2, r2, "\u0394P_i", _fmt(cav.delta_P_incipient, 4), "bar")
        r2 = write_row(ws2, r2, "\u0394P_max",_fmt(cav.delta_P_max, 4),"bar")

    if result.noise:
        noise = result.noise
        ws3 = wb.create_sheet("Noise Analysis")
        ws3.column_dimensions["A"].width = 38
        ws3.column_dimensions["B"].width = 22
        ws3.column_dimensions["C"].width = 22
        r3 = write_section(ws3, 1, "NOISE PREDICTION (IEC 60534-8-3/8-4)")
        r3 = write_row(ws3, r3, "Lpe", f"{noise.Lpe_dba:.1f}", "dBA at 1 m")
        r3 = write_row(ws3, r3, "LWi", f"{noise.LWi_db:.1f}",  "dB re 1 pW")
        r3 = write_row(ws3, r3, "TL",  f"{noise.TL_db:.1f}",   "dB")
        r3 = write_row(ws3, r3, "f_p", f"{noise.f_peak_hz:.0f}", "Hz")
        r3 = write_row(ws3, r3, "eta", f"{noise.eta:.2e}",       "\u2014")
        r3 = write_row(ws3, r3, "Regime", noise.regime,          "")

    if result.messages:
        ws4 = wb.create_sheet("Warnings")
        ws4.column_dimensions["A"].width = 12
        ws4.column_dimensions["B"].width = 30
        ws4.column_dimensions["C"].width = 80
        r4 = write_section(ws4, 1, "ENGINEERING MESSAGES")
        for msg in result.messages:
            ws4.cell(r4, 1, msg.level.value.upper())
            ws4.cell(r4, 2, msg.code)
            ws4.cell(r4, 3, msg.message)
            r4 += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# STREAMLIT RENDER PANEL
# ---------------------------------------------------------------------------

def render_report_panel(result: SizingResult) -> None:
    """Render the report download panel inside the Report tab."""
    from frontend.ui_styles import section_header_html

    if not result.success:
        st.info("Run a successful calculation to generate a report.")
        return

    st.markdown(section_header_html("Export Sizing Report"), unsafe_allow_html=True)

    col1, col2 = st.columns(2)

    with col1:
        st.markdown("#### PDF Report")
        st.caption(
            "Engineering sizing report with data tables and embedded charts "
            "(Cv curve, sizing ratio gauge, pressure profile)."
        )
        try:
            pdf_bytes = build_pdf_bytes(result)
            st.download_button(
                label="Download PDF Report",
                data=pdf_bytes,
                file_name=f"valve_sizing_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
                mime="application/pdf",
                use_container_width=True,
            )
        except ImportError as exc:
            st.warning(f"PDF export unavailable: {exc}")
        except Exception as exc:
            st.error(f"PDF generation failed: {exc}")

    with col2:
        st.markdown("#### Excel Workbook")
        st.caption(
            "Structured Excel workbook with separate sheets for sizing "
            "results, noise analysis, and engineering warnings."
        )
        try:
            xlsx_bytes = build_excel_bytes(result)
            st.download_button(
                label="Download Excel Workbook",
                data=xlsx_bytes,
                file_name=f"valve_sizing_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime=(
                    "application/vnd.openxmlformats-officedocument"
                    ".spreadsheetml.sheet"
                ),
                use_container_width=True,
            )
        except ImportError as exc:
            st.warning(f"Excel export unavailable: {exc}")
        except Exception as exc:
            st.error(f"Excel generation failed: {exc}")

    st.divider()
    st.markdown(section_header_html("Report Contents"), unsafe_allow_html=True)
    st.markdown("""
| Section | Content |
|---|---|
| **1. Primary Results** | Cv required, Cv design, Kv, sizing ratio, opening %, velocity |
| **2. Process Conditions** | P1, P2, T1, flow, density (SI internal) |
| **3. Intermediate Values** | FF, Fp, Y, Fgamma, x, Rev, FR, dP_max, dP_eff |
| **4. Cavitation Analysis** | Regime, sigma, Pvc, dP_incipient, dP_max (liquid only) |
| **5. Noise Prediction** | Lpe, LWi, TL, f_peak, eta, regime |
| **6. Engineering Warnings** | All constraint messages with codes and severity |
| **Charts (page 2)** | Cv characteristic curve, sizing ratio gauge, pressure profile |
    """)
